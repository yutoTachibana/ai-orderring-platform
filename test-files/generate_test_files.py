"""テスト用ファイル生成スクリプト

3種類のテストファイルを生成:
1. 発注仕様書Excel (Slack投稿用)
2. 請求書PDF (請求管理画面でインポート用)
3. 入金CSV (入金消込画面でインポート用)
"""
import os
from datetime import date, datetime
from pathlib import Path

# --- 1. 発注仕様書Excel ---
def create_order_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "発注仕様書"

    # スタイル定義
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # タイトル
    ws.merge_cells("A1:D1")
    ws["A1"] = "発注仕様書"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"作成日: {date.today().strftime('%Y年%m月%d日')}"
    ws["A2"].alignment = Alignment(horizontal="right")

    # 幅設定
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 45

    # --- 発注情報セクション ---
    row = 4
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "■ 発注情報"
    ws[f"A{row}"].font = header_font

    items = [
        ("発注番号", "PO-2026-0215"),
        ("発注日", "2026/02/10"),
        ("発注元企業", "株式会社テックソリューション"),
        ("担当者", "田中太郎"),
    ]
    row = 5
    for label, value in items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border
        ws[f"B{row}"] = value
        ws[f"B{row}"].border = border
        row += 1

    # --- 案件情報セクション ---
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "■ 案件情報"
    ws[f"A{row}"].font = header_font

    row += 1
    project_items = [
        ("案件名", "ECサイトリニューアル開発"),
        ("業務内容", "既存ECサイトのフルリニューアル。React+Next.jsによるフロントエンド刷新、APIのGo移行、インフラのAWS移行を含む。"),
        ("開始日", "2026/04/01"),
        ("終了日", "2026/09/30"),
        ("勤務地", "東京都港区（リモート併用可）"),
        ("契約形態", "準委任"),
    ]
    for label, value in project_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border
        ws[f"B{row}"] = value
        ws[f"B{row}"].border = border
        ws[f"B{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    # --- 条件セクション ---
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "■ 契約条件"
    ws[f"A{row}"].font = header_font

    row += 1
    condition_items = [
        ("月額単価", "¥800,000"),
        ("精算幅下限（H）", "140"),
        ("精算幅上限（H）", "180"),
        ("想定工数", "160時間/月"),
        ("予算", "¥4,800,000"),
        ("必要人数", "1"),
    ]
    for label, value in condition_items:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border
        ws[f"B{row}"] = value
        ws[f"B{row}"].border = border
        row += 1

    # --- スキル要件 ---
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "■ スキル要件"
    ws[f"A{row}"].font = header_font

    row += 1
    ws[f"A{row}"] = "必須スキル"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = border
    ws[f"B{row}"] = "React, TypeScript, Next.js, Go, AWS"
    ws[f"B{row}"].border = border

    row += 1
    ws[f"A{row}"] = "備考"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"A{row}"].border = border
    ws[f"B{row}"] = "フロントエンド経験3年以上、チームリード経験あれば尚可"
    ws[f"B{row}"].border = border
    ws[f"B{row}"].alignment = Alignment(wrap_text=True)

    path = Path("test-files/発注仕様書_テスト.xlsx")
    wb.save(str(path))
    print(f"  [1] {path}")
    return path


# --- 2. 請求書PDF ---
def create_invoice_pdf():
    """fpdf2がなくてもテキストベースで請求書PDFを生成"""
    try:
        from fpdf import FPDF
        _create_invoice_pdf_fpdf()
    except ImportError:
        _create_invoice_pdf_reportlab()


def _create_invoice_pdf_reportlab():
    """reportlabでPDFを生成（フォールバック）"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont

        path = Path("test-files/請求書_テスト.pdf")
        c = canvas.Canvas(str(path), pagesize=A4)
        w, h = A4

        # 日本語フォント
        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
        font = "HeiseiKakuGo-W5"

        # ヘッダー
        c.setFont(font, 24)
        c.drawString(220, h - 60, "請求書")

        c.setFont(font, 10)
        c.drawString(400, h - 100, f"請求日: 2026年02月01日")
        c.drawString(400, h - 115, f"請求番号: INV-2026-0201")

        # 宛先
        c.setFont(font, 12)
        c.drawString(50, h - 100, "株式会社テックソリューション 御中")

        # 発行者
        c.setFont(font, 10)
        c.drawString(400, h - 145, "請求元: 株式会社エスイーエス")
        c.drawString(400, h - 160, "東京都渋谷区1-2-3")
        c.drawString(400, h - 175, "TEL: 03-1234-5678")

        # 対象期間
        c.setFont(font, 10)
        c.drawString(50, h - 160, "対象月: 2026年01月")
        c.drawString(50, h - 175, "支払期日: 2026年02月28日")

        # 請求金額
        c.setFont(font, 16)
        c.drawString(50, h - 210, "ご請求金額: ¥880,000")

        # 明細テーブル
        y = h - 260
        c.setFont(font, 9)
        headers = ["項目", "数量", "単価", "金額"]
        x_positions = [50, 280, 370, 460]
        for i, header in enumerate(headers):
            c.drawString(x_positions[i], y, header)
        c.line(50, y - 5, 550, y - 5)

        items = [
            ("ECサイトリニューアル開発 / SE作業", "160H", "¥5,000", "¥800,000"),
            ("稼働時間: 160.0時間", "", "", ""),
        ]
        y -= 20
        for item in items:
            for i, val in enumerate(item):
                c.drawString(x_positions[i], y, val)
            y -= 18

        # 合計
        y -= 10
        c.line(50, y + 5, 550, y + 5)
        c.drawString(370, y - 10, "小計: ¥800,000")
        c.drawString(370, y - 28, "消費税: ¥80,000")
        c.setFont(font, 12)
        c.drawString(370, y - 50, "合計金額: ¥880,000")

        # 振込先
        y -= 80
        c.setFont(font, 9)
        c.drawString(50, y, "【振込先】")
        c.drawString(50, y - 15, "三菱UFJ銀行 渋谷支店 普通 1234567")
        c.drawString(50, y - 30, "カ）エスイーエス")

        c.save()
        print(f"  [2] {path}")
        return path

    except ImportError:
        # reportlabもない場合: pdfplumberが読めるテキストPDFをopenpyxlの代わりに簡易生成
        _create_invoice_pdf_minimal()


def _create_invoice_pdf_minimal():
    """最小限のPDF生成（外部ライブラリ不要）"""
    path = Path("test-files/請求書_テスト.pdf")

    # 最小限のPDFを手動構築
    content = """%PDF-1.4
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R] /Count 1 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792]
   /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>
endobj
5 0 obj
<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>
endobj
4 0 obj
<< /Length 400 >>
stream
BT
/F1 20 Tf
200 750 Td
(Invoice / Seikyusho) Tj
/F1 10 Tf
0 -30 Td
(Invoice No: INV-2026-0201) Tj
0 -15 Td
(Billing Date: 2026/02/01) Tj
0 -15 Td
(Due Date: 2026/02/28) Tj
0 -15 Td
(Billing Month: 2026/01) Tj
0 -30 Td
(Subtotal: 800,000) Tj
0 -15 Td
(Tax: 80,000) Tj
0 -15 Td
(Total: 880,000) Tj
0 -15 Td
(Working Hours: 160.0H) Tj
ET
endstream
endobj
xref
0 6
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
0000000314 00000 n
0000000266 00000 n
trailer
<< /Root 1 0 R /Size 6 >>
startxref
766
%%EOF
"""
    with open(path, "w") as f:
        f.write(content)
    print(f"  [2] {path} (minimal PDF - ASCII only)")
    return path


def _create_invoice_pdf_fpdf():
    """fpdf2でPDFを生成"""
    from fpdf import FPDF

    path = Path("test-files/請求書_テスト.pdf")
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 24)
    pdf.cell(0, 20, "Seikyusho (Invoice)", align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 10)
    lines = [
        "Invoice No: INV-2026-0201",
        "Billing Date: 2026/02/01",
        "Target Month: 2026/01",
        "Due Date: 2026/02/28",
        "",
        "Subtotal: 800,000",
        "Tax Amount: 80,000",
        "Total Amount: 880,000",
        "Working Hours: 160.0H",
    ]
    for line in lines:
        pdf.cell(0, 8, line, new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(path))
    print(f"  [2] {path}")
    return path


# --- 3. 入金CSV ---
def create_bank_csv():
    """銀行入金明細CSV"""
    path = Path("test-files/入金明細_テスト.csv")
    lines = [
        "入金日,金額,振込人,参照番号,銀行名",
        "2026/02/05,880000,カ）テックソリユーシヨン,INV-2026-0201,三菱UFJ銀行",
        "2026/02/06,550000,カ）サンプルシヨウジ,INV-2026-0115,みずほ銀行",
        "2026/02/07,1200000,カ）グローバルテック,,三井住友銀行",
        "2026/02/08,330000,タナカ タロウ,REF-20260208,ゆうちょ銀行",
        "2026/02/10,750000,カ）エービーシーシステムズ,INV-2026-0180,りそな銀行",
    ]
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    print(f"  [3] {path}")
    return path


# --- 4. 一覧表Excel (複数行パース用) ---
def create_order_list_excel():
    """複数案件が入った発注一覧Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "発注一覧"

    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["発注番号", "案件名", "発注元企業", "エンジニア名", "月額単価", "開始日", "終了日", "必須スキル", "備考"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    data = [
        ["PO-2026-001", "決済システム改修", "株式会社ペイメントテック", "山田花子", 750000, "2026/04/01", "2026/09/30", "Java, Spring Boot, AWS", "経験5年以上"],
        ["PO-2026-002", "社内DXプラットフォーム構築", "株式会社デジタルワークス", "鈴木一郎", 850000, "2026/05/01", "2026/10/31", "Python, React, GCP", "PM経験あれば尚可"],
        ["PO-2026-003", "AI チャットボット開発", "株式会社AIラボ", "佐藤次郎", 900000, "2026/04/15", "2026/12/31", "Python, LLM, FastAPI", "機械学習経験必須"],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # 列幅調整
    widths = [15, 28, 25, 15, 15, 14, 14, 30, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    path = Path("test-files/発注一覧_テスト.xlsx")
    wb.save(str(path))
    print(f"  [4] {path}")
    return path


if __name__ == "__main__":
    os.chdir(Path(__file__).parent.parent)
    print("テストファイルを生成中...\n")
    create_order_excel()
    create_invoice_pdf()
    create_bank_csv()
    create_order_list_excel()
    print("\n完了! test-files/ ディレクトリにファイルが生成されました。")
    print("\n使い方:")
    print("  [1] 発注仕様書_テスト.xlsx → Slackチャンネルに投稿 → 自動パイプライン起動")
    print("  [2] 請求書_テスト.pdf       → 請求管理画面の「PDFインポート」ボタンからアップロード")
    print("  [3] 入金明細_テスト.csv     → 入金消込画面の「CSVインポート」ボタンからアップロード")
    print("  [4] 発注一覧_テスト.xlsx    → 複数案件の一括パーステスト用")
