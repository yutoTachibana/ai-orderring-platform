"""レポート自動生成: 週次/月次レポートをExcel/PDF形式で出力"""
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import SessionLocal
from app.models import Project, Engineer, Contract, Invoice, ProcessingJob


class ReportGenerator:
    """各種レポートを生成する"""

    def generate_monthly_summary(self, year: int, month: int, db=None) -> bytes:
        """月次サマリーレポートをExcelで生成"""
        own_session = db is None
        if own_session:
            db = SessionLocal()
        try:
            wb = Workbook()

            # 案件サマリーシート
            ws = wb.active
            ws.title = "案件サマリー"
            self._write_project_summary(ws, db, year, month)

            # 契約・請求シート
            ws2 = wb.create_sheet("契約・請求")
            self._write_contract_summary(ws2, db, year, month)

            # ジョブ処理実績シート
            ws3 = wb.create_sheet("処理実績")
            self._write_job_summary(ws3, db, year, month)

            output = BytesIO()
            wb.save(output)
            return output.getvalue()
        finally:
            if own_session:
                db.close()

    def _write_project_summary(self, ws, db, year: int, month: int):
        header_font = Font(bold=True, size=12)
        ws.append(["案件サマリーレポート", f"{year}年{month}月"])
        ws["A1"].font = header_font
        ws.append([])
        ws.append(["案件名", "クライアント", "ステータス", "予算", "開始日", "終了日"])

        projects = db.query(Project).all()
        for p in projects:
            ws.append([
                p.name,
                p.client_company.name if p.client_company else "",
                p.status.value if p.status else "",
                p.budget,
                str(p.start_date) if p.start_date else "",
                str(p.end_date) if p.end_date else "",
            ])

    def _write_contract_summary(self, ws, db, year: int, month: int):
        ws.append(["契約・請求サマリー", f"{year}年{month}月"])
        ws.append([])
        ws.append(["契約番号", "エンジニア", "月額", "ステータス", "請求額", "入金状況"])

        contracts = db.query(Contract).filter(Contract.status.in_(["active", "expired"])).all()
        for c in contracts:
            invoice = db.query(Invoice).filter(
                Invoice.contract_id == c.id,
            ).order_by(Invoice.billing_month.desc()).first()
            ws.append([
                c.contract_number,
                c.engineer.full_name if c.engineer else "",
                c.monthly_rate,
                c.status.value if c.status else "",
                invoice.total_amount if invoice else 0,
                invoice.status.value if invoice else "",
            ])

    def _write_job_summary(self, ws, db, year: int, month: int):
        ws.append(["処理ジョブ実績", f"{year}年{month}月"])
        ws.append([])
        ws.append(["ジョブID", "ステータス", "振り分け先", "作成日"])

        jobs = db.query(ProcessingJob).all()
        for j in jobs:
            ws.append([
                j.id,
                j.status.value if j.status else "",
                j.assigned_system or "",
                str(j.created_at) if j.created_at else "",
            ])
