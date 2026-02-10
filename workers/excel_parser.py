"""Excel解析エンジン: openpyxlでExcelをパース、列マッピングに基づいてデータ抽出"""
import json
from pathlib import Path

from openpyxl import load_workbook


class ExcelParseError(Exception):
    pass


class ExcelParser:
    """Excelファイルを解析し、テンプレート定義に基づいてデータを抽出する"""

    def __init__(self, template: dict | None = None):
        self.template = template or {}
        self.column_mappings = template.get("column_mappings", {}) if template else {}
        self.validation_rules = template.get("validation_rules", {}) if template else {}

    def parse(self, file_path: str) -> list[dict]:
        """Excelファイルをパースしてレコードのリストを返す"""
        path = Path(file_path)
        if not path.exists():
            raise ExcelParseError(f"ファイルが見つかりません: {file_path}")

        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ExcelParseError("アクティブなシートが見つかりません")

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise ExcelParseError("データ行が見つかりません")

        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        records = []
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            record = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers):
                    key = self.column_mappings.get(headers[col_idx], headers[col_idx])
                    record[key] = value

            validation_error = self._validate(record, row_idx)
            if validation_error:
                errors.append(validation_error)
                continue

            records.append(record)

        wb.close()
        return records

    def _validate(self, record: dict, row_idx: int) -> str | None:
        """バリデーションルールに基づいて検証"""
        required_fields = self.validation_rules.get("required", [])
        for field in required_fields:
            if not record.get(field):
                return f"行{row_idx}: 必須項目 '{field}' が空です"
        return None

    def parse_order_excel(self, file_path: str) -> dict:
        """発注仕様書Excel専用のパース処理"""
        path = Path(file_path)
        if not path.exists():
            raise ExcelParseError(f"ファイルが見つかりません: {file_path}")

        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ExcelParseError("アクティブなシートが見つかりません")

        data = {}
        for row in ws.iter_rows(min_row=1, max_row=50, max_col=10, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    key = cell.value.strip()
                    # Look for value in next column
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if next_cell.value is not None:
                        data[key] = next_cell.value

        wb.close()
        return data
